from docx import Document
from openpyxl import Workbook, load_workbook
import random
import PyPDF2


def zamena(s=[str(input()), str(input()), str(input())]):
    document = Document(s[0])
    p = document.paragraphs
    for i in p:
        if i.text == s[1]:
            print(i.text)
            i.text = s[2]
    document.save('Pytest.docx')


def randomexcel(name):
    docxl = load_workbook(filename=str(name))
    doclist = docxl['python']
    for i in range(1, 13):
        for j in range(1, 13):
            doclist.cell(column=i, row=j, value=random.randint(1, 25))
    docxl.save(filename=str(name))


def pdfview(name):
    pdf_file = open(str(name), 'rb')
    read_pdf = PyPDF2.PdfFileReader(pdf_file)
    page1 = read_pdf.getPage(0)
    page_content = page1.extractText()
    print(page_content.encode('utf-8'))
